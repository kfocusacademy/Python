# program to read sales data from excel, convert value to INR using the latest currency exchange rate, update the excel sheet
import requests
from openpyxl import load_workbook
import datetime
import os
import time

# Function to get the latest USD to INR exchange rate
def get_exchange_inr(from_currency):
    try:
        url = f"https://open.er-api.com/v6/latest/{from_currency}"
        response = requests.get(url).json()
        value_inr = response['rates'][f'INR']
        return value_inr
    except Exception as e:
        print(e)
        return None
    
# Function to convert sales data to INR and update the Excel file
def convert_sales_to_inr(file_path, sheet_name='sales_data'):
    try:
        # Load the workbook and select the sheet
        workbook = load_workbook(filename=file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            return

        # Iterate through the rows and convert sales to INR
        for row in sheet.iter_rows(min_row=2):  # Assuming sales data starts from row 2
            print(row[5].value)
            if row[5].value == "Yes":  # Assuming column F (index 5) indicates if already converted
                continue
            sales_value = row[2].value  # Assuming sales value is in column c (index 2)
            sales_currency = row[3].value  # Assuming currency code is in column D (index 3)
            # Get the exchange rate
            exchange_rate = get_exchange_inr(sales_currency)
            if exchange_rate is None:
                print("Failed to retrieve exchange rate.")
                return
            
            if isinstance(sales_value, (int, float)):
                sales_in_inr = sales_value * exchange_rate
                row[4].value = round(sales_in_inr, 2)  # Assuming INR value is to be written in column E (index 4)
                row[5].value = "Yes"  # Assuming INR value is to be written in column E (index 4)

        # Save the updated workbook
        workbook.save(filename=file_path)
        print(f"Sales data converted to INR and updated in '{file_path}'.")

    except Exception as e:
        print(f"An error occurred: {e}")
        

# Driver Code        
if __name__ == "__main__":        
    file_path = r"C:\Ravi\Python\Projects\Currency_Capture\sales_data.xlsx"  # Path to your Excel file
    convert_sales_to_inr(file_path)
    